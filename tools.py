"""Tool definitions and implementations for optimus Coding tool."""
import os
import sys
import re
import json
import csv
import tempfile
import subprocess
from pathlib import Path
from typing import Callable, Optional

# ---- Tool JSON schemas ----

TOOL_SCHEMAS = [
    {
        "name": "Read",
        "description": (
            "Read a file's contents. Returns content with line numbers "
            "(format: 'N\\tline'). Use limit/offset to read large files in chunks."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string", "description": "Absolute file path"},
                "limit":     {"type": "integer", "description": "Max lines to read"},
                "offset":    {"type": "integer", "description": "Start line (0-indexed)"},
            },
            "required": ["file_path"],
        },
    },
    {
        "name": "Write",
        "description": "Write content to a file, creating parent directories as needed.",
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path": {"type": "string"},
                "content":   {"type": "string"},
            },
            "required": ["file_path", "content"],
        },
    },
    {
        "name": "Edit",
        "description": (
            "Replace exact text in a file. old_string must match exactly (including whitespace). "
            "If old_string appears multiple times, use replace_all=true or add more context."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "file_path":   {"type": "string"},
                "old_string":  {"type": "string", "description": "Exact text to replace"},
                "new_string":  {"type": "string", "description": "Replacement text"},
                "replace_all": {"type": "boolean", "description": "Replace all occurrences"},
            },
            "required": ["file_path", "old_string", "new_string"],
        },
    },
    {
        "name": "Bash",
        "description": "Execute a shell command. Returns stdout+stderr. Stateless (no cd persistence).",
        "input_schema": {
            "type": "object",
            "properties": {
                "command": {"type": "string"},
                "timeout": {"type": "integer", "description": "Seconds before timeout (default 30)"},
            },
            "required": ["command"],
        },
    },
    {
        "name": "Glob",
        "description": "Find files matching a glob pattern. Returns sorted list of matching paths.",
        "input_schema": {
            "type": "object",
            "properties": {
                "pattern": {"type": "string", "description": "Glob pattern e.g. **/*.py"},
                "path":    {"type": "string", "description": "Base directory (default: cwd)"},
            },
            "required": ["pattern"],
        },
    },
    {
        "name": "Grep",
        "description": "Search file contents with regex using ripgrep (falls back to grep).",
        "input_schema": {
            "type": "object",
            "properties": {
                "pattern":      {"type": "string", "description": "Regex pattern"},
                "path":         {"type": "string", "description": "File or directory to search"},
                "glob":         {"type": "string", "description": "File filter e.g. *.py"},
                "output_mode":  {
                    "type": "string",
                    "enum": ["content", "files_with_matches", "count"],
                    "description": "content=matching lines, files_with_matches=file paths, count=match counts",
                },
                "case_insensitive": {"type": "boolean"},
                "context":      {"type": "integer", "description": "Lines of context around matches"},
            },
            "required": ["pattern"],
        },
    },
    {
        "name": "WebFetch",
        "description": "Fetch a URL and return its text content (HTML stripped).",
        "input_schema": {
            "type": "object",
            "properties": {
                "url":    {"type": "string"},
                "prompt": {"type": "string", "description": "Hint for what to extract"},
            },
            "required": ["url"],
        },
    },
    {
        "name": "WebSearch",
        "description": "Search the web via DuckDuckGo and return top results.",
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string"},
            },
            "required": ["query"],
        },
    },
    {
        "name": "ExcelAutomate",
        "description": (
            "Automate Excel operations based on natural language requests. "
            "Can read, write, format data, create charts. "
            "If no file path is provided, creates a new workbook. "
            "Always defaults to Sheet1 unless user specifies a different sheet_name. "
            "If file_path is provided but the file does not exist, creates a new workbook and saves it there."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "request": {
                    "type": "string",
                    "description": "Natural language description of what to do in Excel"
                },
                "file_path": {
                    "type": "string",
                    "description": "Optional path to Excel file. If file exists it is opened; if not, a new workbook is created and saved to this path."
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Sheet name. Always defaults to 'Sheet1' when not specified."
                },
                "output_format": {
                    "type": "string",
                    "enum": ["json", "text", "file_path"],
                    "description": "How to return results: 'json', 'text', or 'file_path'"
                },
                "data": {
                    "type": "array",
                    "items": {"type": "array"},
                    "description": (
                        "Optional structured data to write. A list of rows, "
                        "where each row is a list of cell values. "
                        "The first row should be the header. "
                        "Example: [['Year','Price'],[2020,41],[2021,68]]. "
                        "When writing data, ALWAYS provide this parameter with the actual data "
                        "instead of embedding data in the request text."
                    )
                }
            },
            "required": ["request"],
        },
    },
    {
        "name": "EmailSend",
        "description": (
            "Send an email via the SMTP2Go API. "
            "Only recipient email(s) 'to' is required - all other fields (subject, sender, body, cc, bcc) "
            "are optional and will be auto-generated in a professional tone based on context if omitted."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "sender": {
                    "type": "string",
                    "description": "Sender email address (defaults to marva112@suchance.com if not specified)"
                },
                "subject": {
                    "type": "string",
                    "description": "Email subject line (auto-generated in a professional tone if not specified)"
                },
                "to": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of recipient email addresses (required)"
                },
                "cc": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Optional CC recipients"
                },
                "bcc": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Optional BCC recipients"
                },
                "text_body": {
                    "type": "string",
                    "description": "Plain-text email body (auto-generated in a professional tone if not specified)"
                },
                "html_body": {
                    "type": "string",
                    "description": "HTML email body (optional)"
                },
            },
            "required": ["to"],
        },
    },
]

# ---- Safe bash ----

_SAFE_PREFIXES = (
    "ls", "cat", "head", "tail", "wc", "pwd", "echo", "printf", "date",
    "which", "type", "env", "printenv", "uname", "whoami", "id",
    "git log", "git status", "git diff", "git show", "git branch",
    "git remote", "git stash list", "git tag",
    "find ", "grep ", "rg ", "ag ", "fd ",
    "python ", "python3 ", "node ", "ruby ", "perl ",
    "pip show", "pip list", "npm list", "cargo metadata",
    "df ", "du ", "free ", "top -bn", "ps ",
    "curl -I", "curl --head",
)

def _is_safe_bash(cmd):
    c = cmd.strip()
    return any(c.startswith(p) for p in _SAFE_PREFIXES)

# ---- Tool implementations ----

def _read(file_path, limit=None, offset=None):
    p = Path(file_path)
    if not p.exists():
        return f"Error: file not found: {file_path}"
    if p.is_dir():
        return f"Error: {file_path} is a directory"
    try:
        lines = p.read_text(errors="replace").splitlines(keepends=True)
        start = offset or 0
        chunk = lines[start:start + limit] if limit else lines[start:]
        if not chunk:
            return "(empty file)"
        return "".join(f"{start + i + 1}\t{l}" for i, l in enumerate(chunk))
    except Exception as e:
        return f"Error: {e}"


def _write(file_path, content):
    p = Path(file_path)
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(content)
        lc = content.count("\n") + (1 if content and not content.endswith("\n") else 0)
        return f"Wrote {lc} lines to {file_path}"
    except Exception as e:
        return f"Error: {e}"


def _edit(file_path, old_string, new_string, replace_all=False):
    p = Path(file_path)
    if not p.exists():
        return f"Error: file not found: {file_path}"
    try:
        content = p.read_text()
        count = content.count(old_string)
        if count == 0:
            return "Error: old_string not found in file"
        if count > 1 and not replace_all:
            return (f"Error: old_string appears {count} times. "
                    "Provide more context to make it unique, or use replace_all=true.")
        new_content = content.replace(old_string, new_string, count if replace_all else 1)
        p.write_text(new_content)
        return f"Replaced {'all ' + str(count) if replace_all else '1'} occurrence(s) in {file_path}"
    except Exception as e:
        return f"Error: {e}"


def _bash(command, timeout=30):
    try:
        r = subprocess.run(
            command, shell=True, capture_output=True, text=True,
            timeout=timeout, cwd=os.getcwd(),
        )
        out = r.stdout
        if r.stderr:
            out += ("\n" if out else "") + "[stderr]\n" + r.stderr
        return out.strip() or "(no output)"
    except subprocess.TimeoutExpired:
        return f"Error: timed out after {timeout}s"
    except Exception as e:
        return f"Error: {e}"


def _glob(pattern, path=None):
    base = Path(path) if path else Path.cwd()
    try:
        matches = sorted(base.glob(pattern))
        if not matches:
            return "No files matched"
        return "\n".join(str(m) for m in matches[:500])
    except Exception as e:
        return f"Error: {e}"


def _has_rg():
    try:
        subprocess.run(["rg", "--version"], capture_output=True, check=True)
        return True
    except Exception:
        return False


def _grep(pattern, path=None, glob=None,
          output_mode="files_with_matches",
          case_insensitive=False, context=0):
    use_rg = _has_rg()
    cmd = ["rg" if use_rg else "grep", "--no-heading"]
    if case_insensitive:
        cmd.append("-i")
    if output_mode == "files_with_matches":
        cmd.append("-l")
    elif output_mode == "count":
        cmd.append("-c")
    else:
        cmd.append("-n")
        if context:
            cmd += ["-C", str(context)]
    if glob:
        cmd += (["--glob", glob] if use_rg else ["--include", glob])
    cmd.append(pattern)
    cmd.append(path or str(Path.cwd()))
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        out = r.stdout.strip()
        return out[:20000] if out else "No matches found"
    except Exception as e:
        return f"Error: {e}"


def _webfetch(url, prompt=None):
    try:
        import httpx
        r = httpx.get(url, headers={"User-Agent": "NanoClaude/1.0"},
                      timeout=30, follow_redirects=True)
        r.raise_for_status()
        ct = r.headers.get("content-type", "")
        if "html" in ct:
            text = re.sub(r"<script[^>]*>.*?</script>", "", r.text,
                          flags=re.DOTALL | re.IGNORECASE)
            text = re.sub(r"<style[^>]*>.*?</style>", "", text,
                          flags=re.DOTALL | re.IGNORECASE)
            text = re.sub(r"<[^>]+>", " ", text)
            text = re.sub(r"\s+", " ", text).strip()
        else:
            text = r.text
        return text[:25000]
    except ImportError:
        return "Error: httpx not installed -- run: pip install httpx"
    except Exception as e:
        return f"Error: {e}"


def _websearch(query):
    try:
        import httpx
        url = "https://html.duckduckgo.com/html/"
        r = httpx.get(url, params={"q": query},
                      headers={"User-Agent": "Mozilla/5.0 (compatible)"},
                      timeout=30, follow_redirects=True)
        titles = re.findall(r'class="result__title"[^>]*>.*?<a[^>]*href="([^"]+)"[^>]*>(.*?)</a>',
                            r.text, re.DOTALL)
        snippets = re.findall(r'class="result__snippet"[^>]*>(.*?)</div>', r.text, re.DOTALL)
        results = []
        for i, (link, title) in enumerate(titles[:8]):
            t = re.sub(r"<[^>]+>", "", title).strip()
            s = re.sub(r"<[^>]+>", "", snippets[i]).strip() if i < len(snippets) else ""
            results.append(f"**{t}**\n{link}\n{s}")
        return "\n\n".join(results) if results else "No results found"
    except ImportError:
        return "Error: httpx not installed -- run: pip install httpx"
    except Exception as e:
        return f"Error: {e}"


# ---- Data Normalization Helper ----

def _normalize_excel_data(raw):
    """Convert any common JSON shape into a list-of-rows (header row first).

    Handles:
    - list of lists (already correct)  [['Name','Age'],['John',30]]
    - list of dicts                    [{'Name':'John','Age':30}]
    - dict of lists                    {'Name':['John','Jane'],'Age':[30,25]}
    - dict of dicts                    {'r1':{'Name':'John','Age':30}}
    - plain dict (single row)          {'Name':'John','Age':30}
    - stringified versions of all above (recursive)
    - strings containing JSON text embedded in a cell
    """
    if raw is None:
        return None

    # If it's already a list-of-lists, validate and return
    if isinstance(raw, list) and all(isinstance(r, (list, tuple)) for r in raw):
        return [list(r) for r in raw]

    if isinstance(raw, str):
        # Try to parse as JSON first
        try:
            parsed = json.loads(raw)
            return _normalize_excel_data(parsed)
        except (json.JSONDecodeError, TypeError):
            return None

    # List of dicts
    if isinstance(raw, list) and len(raw) > 0 and isinstance(raw[0], dict):
        # Collect all keys preserving order
        seen_keys = []
        for item in raw:
            for k in item.keys():
                if k not in seen_keys:
                    seen_keys.append(k)
        rows = [seen_keys]
        for item in raw:
            row = [item.get(k) for k in seen_keys]
            rows.append(row)
        return rows

    # Dict of lists  {"Name":["John","Jane"],"Age":[30,25]}
    if isinstance(raw, dict):
        vals = list(raw.values())
        if len(vals) > 0 and all(isinstance(v, list) for v in vals):
            keys = list(raw.keys())
            max_len = max(len(v) for v in vals) if vals else 0
            rows = [keys]
            for i in range(max_len):
                row = [vals[j][i] if i < len(vals[j]) else None for j in range(len(keys))]
                rows.append(row)
            return rows

        # Dict of dicts  {"r1":{"Name":"John","Age":30},"r2":{"Name":"Jane","Age":25}}
        if len(vals) > 0 and all(isinstance(v, dict) for v in vals):
            seen_keys = []
            for v in vals:
                for k in v.keys():
                    if k not in seen_keys:
                        seen_keys.append(k)
            rows = [seen_keys]
            for v in vals:
                rows.append([v.get(k) for k in seen_keys])
            return rows

        # Plain single-row dict {"Name":"John","Age":30}
        return [list(raw.keys()), list(raw.values())]

    # Single value
    return [[raw]]


# ---- ExcelAutomate ----
# Key rules:
#   1. sheet_name ALWAYS defaults to "Sheet1" unless user specifies otherwise.
#   2. If file_path is provided but file does NOT exist, a new workbook is
#      created and saved to that path (instead of a temp file).
#   3. The parent directory of file_path is created if missing.

def _excel_automate(request, file_path=None, sheet_name=None, output_format="text", data=None):
    """Execute Excel automation based on natural language request and optional structured data."""
    try:
        is_windows = sys.platform == "win32"

        try:
            import openpyxl
            from openpyxl.utils import get_column_letter, column_index_from_string
        except ImportError:
            return "Error: openpyxl not installed -- run: pip install openpyxl"

        excel_com_available = False
        if is_windows:
            try:
                import win32com.client
                excel_com_available = True
            except ImportError:
                pass

        # ALWAYS default to Sheet1
        if not sheet_name:
            sheet_name = "Sheet1"

        request_lower = request.lower()

        # ---- Open existing or create new ----
        # If file_path given and file exists -> open it
        # Otherwise -> create a new workbook (save it at file_path later)
        if file_path and Path(file_path).exists():
            try:
                wb = openpyxl.load_workbook(file_path)
            except Exception as e:
                return f"Error opening workbook: {e}"
        else:
            wb = openpyxl.Workbook()
            # rename whatever openpyxl created as default to Sheet1
            if "Sheet" in wb.sheetnames:
                wb["Sheet"].title = "Sheet1"

        # ---- Select or create target sheet ----
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)

        result = None

        # ---- Sheet management ----
        if "remove other sheets" in request_lower or "delete all but" in request_lower:
            sheets_to_remove = [s for s in wb.sheetnames if s != sheet_name]
            for s_name in sheets_to_remove:
                del wb[s_name]
            result = f"Removed sheets: {', '.join(sheets_to_remove) if sheets_to_remove else 'None'}"

        # ---- Read ----
        elif "read" in request_lower or "get" in request_lower:
            data = [[cell.value for cell in row] for row in ws.iter_rows()]
            if output_format == "json":
                return json.dumps(data, indent=2)
            if output_format == "file_path":
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv",
                                                  mode="w", newline="")
                csv.writer(tmp).writerows(data)
                tmp.close()
                return tmp.name
            result = "\n".join(
                ",".join(str(c) if c is not None else "" for c in row)
                for row in data
            )

        # ---- Write / create / populate ----
        elif any(kw in request_lower for kw in (
                "write", "set", "put", "create", "add", "populate",
                "save", "data", "table", "list")) or "data" in request_lower:

            # Priority 1: structured data parameter passed directly
            # Priority 2: JSON array/dict embedded in request text (regex fallback)
            parsed_data = None

            if data is not None:
                # data can be a python list/dict (direct arg) or a JSON string
                parsed_data = data

            if parsed_data is None:
                # Try to extract and parse JSON from request text
                json_match = re.search(r'[\[{]\s*.*?[\]}]\s*', request, re.DOTALL)
                if json_match:
                    try:
                        parsed_data = json.loads(json_match.group(0))
                    except (json.JSONDecodeError, Exception):
                        # If first bracket pair fails, try to find the longest valid JSON
                        for end_pos in range(len(request), len(json_match.group(0)), -1):
                            substring = request[len(json_match.group(0)) * 0:end_pos]
                            bracket_match = re.search(r'(\[.*\]|\{.*\})', request[:end_pos], re.DOTALL)
                            if bracket_match:
                                try:
                                    parsed_data = json.loads(bracket_match.group(0))
                                    break
                                except (json.JSONDecodeError, Exception):
                                    pass

            if parsed_data is not None:
                parsed_data = _normalize_excel_data(parsed_data)

            if parsed_data is None:
                # Fallback: try to extract tabular-like text from the request
                # Look for lines with commas/tabs as separators
                lines = request.split('\n')
                table_lines = [line.strip() for line in lines if line.strip() and '|' not in line]
                if len(table_lines) >= 2:
                    parsed_data = []
                    for line in table_lines:
                        if ',' in line:
                            parsed_data.append([v.strip() for v in line.split(',')])
                        elif '\t' in line:
                            parsed_data.append([v.strip() for v in line.split('\t')])
                        else:
                            parsed_data.append([line.strip()])
                    if parsed_data and all(len(r) == len(parsed_data[0]) for r in parsed_data):
                        pass  # valid table
                    else:
                        parsed_data = None

            if parsed_data is None:
                return ("Error: no data found. Provide data as a list of rows via the 'data' "
                        "parameter, e.g. data=[['Year','Revenue'],[2020,100]], "
                        "or as a list of dictionaries, or embed JSON in the request.")

            for i, row in enumerate(parsed_data):
                vals = list(row) if isinstance(row, (list, tuple)) else [row]
                for j, v in enumerate(vals):
                    # Recursively unwrap any still-stringified JSON in cells
                    if isinstance(v, str):
                        try:
                            unwrapped = json.loads(v)
                            v = unwrapped
                        except (json.JSONDecodeError, TypeError):
                            pass
                    # Try numeric conversion
                    if isinstance(v, str):
                        try:
                            v = float(v)
                            if v == int(v):
                                v = int(v)
                        except (ValueError, TypeError):
                            pass
                    # Flatten nested structures that slipped through
                    if isinstance(v, (dict, list)):
                        v = json.dumps(v)
                    ws.cell(row=i + 1, column=j + 1, value=v)
            result = f"Wrote {len(parsed_data)} rows to '{sheet_name}'"

        # ---- Format ----
        elif "format" in request_lower or "style" in request_lower:
            if "header" in request_lower or "first row" in request_lower:
                for cell in ws[1]:
                    cell.font = openpyxl.styles.Font(bold=True)
                result = f"Bolded header row in '{sheet_name}'"
            elif "bold" in request_lower:
                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = openpyxl.styles.Font(bold=True)
                result = f"Applied bold formatting to '{sheet_name}'"
            elif "italic" in request_lower:
                for row in ws.iter_rows():
                    for cell in row:
                        cell.font = openpyxl.styles.Font(italic=True)
                result = f"Applied italic formatting to '{sheet_name}'"
            elif "background" in request_lower or "fill" in request_lower:
                for row in ws.iter_rows():
                    for cell in row:
                        cell.fill = openpyxl.styles.PatternFill(
                            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                        )
                result = f"Applied yellow fill to '{sheet_name}'"
            else:
                result = "Formatting applied"

        # ---- Sum / Total ----
        elif "sum" in request_lower or "total" in request_lower:
            col_m = re.search(r'column\s+([A-Z]+)', request, re.IGNORECASE)
            if col_m:
                ci = column_index_from_string(col_m.group(1).upper())
                total = sum(
                    c.value
                    for row in ws.iter_rows(min_col=ci, max_col=ci)
                    for c in row
                    if isinstance(c.value, (int, float))
                )
                result = f"Sum of column {col_m.group(1)}: {total}"
            else:
                range_m = re.search(r'range\s+([A-Z]+\d*:[A-Z]+\d*)', request, re.IGNORECASE)
                if range_m:
                    try:
                        cells = ws[range_m.group(1)]
                        total = sum(
                            c.value for row in cells for c in row
                            if isinstance(c.value, (int, float))
                        )
                        result = f"Sum of range {range_m.group(1)}: {total}"
                    except Exception as e:
                        return f"Error calculating sum: {e}"
                else:
                    return "Error: no column or range specified for sum."

        # ---- Chart ----
        elif "chart" in request_lower or "graph" in request_lower:
            if not is_windows or not excel_com_available:
                return "Chart creation requires Windows + Excel COM (pip install pywin32)"
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                tp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
                wb.save(tp)
                book = excel.Workbooks.Open(tp)
                cs = book.Sheets.Add()
                cs.Name = "Chart"
                chart = cs.Shapes.AddChart2(240, 51).Chart
                ds = book.Sheets[0]
                lr = ds.UsedRange.Rows.Count
                lc = ds.UsedRange.Columns.Count
                if lr > 1 and lc >= 2:
                    letters = ""
                    tmp_c = lc
                    while tmp_c > 0:
                        tmp_c, rem = divmod(tmp_c - 1, 26)
                        letters = chr(65 + rem) + letters
                    chart.SetSourceData(ds.Range(f"A1:{letters}{lr}"))
                else:
                    chart.SetSourceData(ds.Range("A1:B10"))
                book.Save()
                excel.Application.Quit()
                result = f"Chart created in {tp}"
            except Exception as e:
                return f"Error creating chart: {e}"

        # ---- Default ----
        if result is None:
            result = "Operation completed"

        # ---- Save workbook ----
        if file_path:
            try:
                # Create parent dirs if missing
                Path(file_path).parent.mkdir(parents=True, exist_ok=True)
                wb.save(file_path)
                if output_format == "file_path":
                    return file_path
            except Exception as e:
                return f"Error saving workbook to {file_path}: {e}"
        else:
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            wb.save(tmp.name)
            tmp.close()
            if output_format == "file_path":
                return tmp.name

        return result

    except Exception as e:
        return f"Error in ExcelAutomate: {e}"


def _email_send(sender, subject, **inputs):
    """Send an email using the SMTP2Go API."""
    try:
        from config import CONFIG_FILE
        cfg = json.loads(CONFIG_FILE.read_text()) if CONFIG_FILE.exists() else {}
        api_key = cfg.get("smtp_api_key", os.environ.get("SMTP_API_KEY", ""))
        if not api_key:
            return "Error: smtp_api_key not configured in ~/.optimus/config.json or SMTP_API_KEY environment variable."

        base_url = cfg.get("smtp_base_url", "https://api.smtp2go.com/v3/").rstrip("/")

        # Auto-generate a professional subject if none provided
        if not subject or not subject.strip():
            subject = "Regarding Your Request"

        payload = {
            "api_key": api_key,
            "sender": sender,
            "subject": subject,
        }

        to = inputs.get("to")
        if isinstance(to, list) and len(to) > 0:
            payload["to"] = to
        elif isinstance(to, str):
            payload["to"] = [to]
        else:
            return "Error: 'to' (recipients) is required but was not provided. Please specify at least one recipient."

        cc = inputs.get("cc")
        if isinstance(cc, list) and len(cc) > 0:
            payload["cc"] = cc

        bcc = inputs.get("bcc")
        if isinstance(bcc, list) and len(bcc) > 0:
            payload["bcc"] = bcc

        text_body = inputs.get("text_body")
        if text_body:
            payload["text_body"] = text_body
        else:
            payload["text_body"] = ""

        html_body = inputs.get("html_body")
        if html_body:
            payload["html_body"] = html_body

        import httpx
        r = httpx.post(
            f"{base_url}/email/send",
            json=payload,
            timeout=30,
        )
        data = r.json()

        if r.status_code == 200 and data.get("request_id"):
            return f"Email sent successfully. Request ID: {data['request_id']}"
        else:
            return f"Email sending failed: {json.dumps(data, indent=2)}"
    except ImportError:
        return "Error: httpx not installed -- run: pip install httpx"
    except Exception as e:
        return f"Error in EmailSend: {e}"


# ---- Dispatcher ----

def execute_tool(
    name,
    inputs,
    permission_mode="auto",
    ask_permission=None,
):
    """Dispatch tool execution; ask permission for write/destructive ops."""

    def _check(desc):
        if permission_mode == "accept-all":
            return True
        if ask_permission:
            return ask_permission(desc)
        return True  # headless: allow everything

    if name == "Read":
        return _read(inputs["file_path"], inputs.get("limit"), inputs.get("offset"))

    elif name == "Write":
        if not _check(f"Write to {inputs['file_path']}"):
            return "Denied: user rejected write operation"
        return _write(inputs["file_path"], inputs["content"])

    elif name == "Edit":
        if not _check(f"Edit {inputs['file_path']}"):
            return "Denied: user rejected edit operation"
        return _edit(inputs["file_path"], inputs["old_string"],
                     inputs["new_string"], inputs.get("replace_all", False))

    elif name == "Bash":
        cmd = inputs["command"]
        if permission_mode != "accept-all" and not _is_safe_bash(cmd):
            if not _check(f"Bash: {cmd}"):
                return "Denied: user rejected bash command"
        return _bash(cmd, inputs.get("timeout", 30))

    elif name == "Glob":
        return _glob(inputs["pattern"], inputs.get("path"))

    elif name == "Grep":
        return _grep(
            inputs["pattern"], inputs.get("path"), inputs.get("glob"),
            inputs.get("output_mode", "files_with_matches"),
            inputs.get("case_insensitive", False),
            inputs.get("context", 0),
        )

    elif name == "ExcelAutomate":
        if inputs.get("file_path") or inputs.get("output_format") == "file_path":
            if not _check(f"Excel operation: {inputs['request']}"):
                return "Denied: user rejected Excel operation"
        return _excel_automate(
            inputs["request"],
            inputs.get("file_path"),
            inputs.get("sheet_name", "Sheet1"),
            inputs.get("output_format", "text"),
            inputs.get("data"),  # Pass the data parameter!
        )

    elif name == "WebFetch":
        return _webfetch(inputs["url"], inputs.get("prompt"))

    elif name == "WebSearch":
        return _websearch(inputs["query"])

    elif name == "EmailSend":
        to_val = inputs.get("to")
        if not to_val:
            return "Error: at least one recipient ('to') is required to send an email."
        if not _check(f"Send email to {to_val}"):
            return "Denied: user rejected email send operation"
        sender = inputs.get("sender", "marva112@suchance.com")
        subject = inputs.get("subject", "")
        text_body = inputs.get("text_body", "")
        html_body = inputs.get("html_body", "")
        extra = {}
        if "cc" in inputs:
            extra["cc"] = inputs["cc"]
        if "bcc" in inputs:
            extra["bcc"] = inputs["bcc"]
        return _email_send(sender, subject, to=to_val, text_body=text_body,
                           html_body=html_body, **extra)

    else:
        return f"Unknown tool: {name}"


__all__ = ["execute_tool", "TOOL_SCHEMAS"]
